import openpyxl

class ExcelFun():

    def __init__(self, driver):
        self.driver = driver
    
    #cuenta el numero de filas que tiene nuestra hoja de excel
    def getRowCount(file, path, sheetName):
        worbook = openpyxl.load_workbook(path)
        sheet = worbook[sheetName]
        return (sheet.max_row)
    
    def getColumnCount(file, sheetName):
        worbook = openpyxl.load_workbook(file)
        sheet = worbook[sheetName]
        return (sheet.max_column)
    
    def readData(file, path, sheetName, rowNum, columnNo):
        worbook = openpyxl.load_workbook(path)
        sheet = worbook[sheetName]
        return sheet.cell(row=rowNum, column=columnNo).value
    
    def writeData(file, path, sheetName, rowNum, columnNo, data):
        worbook = openpyxl.load_workbook(path)
        sheet = worbook[sheetName]
        sheet.cell(row=rowNum, column=columnNo).value = data
        worbook.save(path)